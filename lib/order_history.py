#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Amazon の購入履歴情報を取得します．

Usage:
  order_history.py [-c CONFIG] [-o EXCEL]

Options:
  -c CONFIG     : CONFIG を設定ファイルとして読み込んで実行します．[default: config.yaml]
  -o EXCEL      : CONFIG を設定ファイルとして読み込んで実行します．[default: amazhist.xlsx]
"""

import os
import re
import math
import pathlib
import datetime
import io
import random
import logging
import inspect
import time
import traceback

import openpyxl
import openpyxl.utils
import openpyxl.styles
import openpyxl.drawing.image
import openpyxl.drawing.xdr
import openpyxl.drawing.spreadsheet_drawing

import crawl_handle
from store_amazon import gen_order_url

TABLE_HEADER = {
    "row": {
        "pos": 2,
        "height": 80,
    },
    "col": {
        "date": {
            "label": "日付",
            "pos": 2,
            "width": 23,
            "format": 'yyyy"年"mm"月"dd"日 ("aaa")"',
        },
        "name": {
            "label": "商品名",
            "pos": 3,
            "width": 70,
            "wrap": True,
        },
        "image": {
            "label": "画像",
            "pos": 4,
            "width": 12,
        },
        "count": {
            "label": "数量",
            "pos": 5,
            "format": "0_ ",
            "width": 8,
        },
        "price": {
            "label": "価格",
            "pos": 6,
            "width": 16,
            "format": '_ ¥* #,##0_ ;_ ¥* -#,##0_ ;_ ¥* "-"_ ;_ @_ ',  # NOTE: 末尾の空白要
        },
        "category": {
            "label": "カテゴリ",
            "pos": 7,
            "length": 3,
            "width": 20,
            "wrap": True,
        },
        "seller": {
            "label": "売り手",
            "pos": 10,
            "width": 29,
            "wrap": True,
        },
        "asin": {
            "label": "商品ID(ASIN)",
            "pos": 11,
            "width": 17,
            "format": "@",
        },
        "no": {
            "label": "注文番号",
            "pos": 12,
            "width": 28,
            "format": "@",
        },
    },
}
TABLE_SHEET_TITLE = "購入アイテム一覧"


def gen_text_pos(row, col):
    return "{col}{row}".format(
        row=row,
        col=openpyxl.utils.get_column_letter(col),
    )


def set_header_cell_style(sheet, row, col, value, width, style):
    sheet.cell(row, col).value = value
    sheet.cell(row, col).style = "Normal"
    sheet.cell(row, col).border = style["border"]
    sheet.cell(row, col).fill = style["fill"]

    if width is not None:
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width


def set_item_cell_style(sheet, row, col, value, style):
    sheet.cell(row, col).value = value
    sheet.cell(row, col).style = "Normal"
    sheet.cell(row, col).border = style["border"]
    sheet.cell(row, col).alignment = openpyxl.styles.Alignment(wrap_text=style["text_wrap"], vertical="top")

    if "text_format" in style:
        sheet.cell(row, col).number_format = style["text_format"]


def insert_table_header(sheet, row, style):
    crawl_handle.set_status(handle, "テーブルのヘッダを設定しています...")

    for key in TABLE_HEADER["col"].keys():
        col = TABLE_HEADER["col"][key]["pos"]
        if "width" in TABLE_HEADER["col"][key]:
            width = TABLE_HEADER["col"][key]["width"]
        else:
            width = None

        if key == "category":
            for i in range(TABLE_HEADER["col"][key]["length"]):
                set_header_cell_style(
                    sheet,
                    row,
                    col + i,
                    TABLE_HEADER["col"][key]["label"] + " ({i})".format(i=i + 1),
                    width,
                    style,
                )
        else:
            set_header_cell_style(sheet, row, col, TABLE_HEADER["col"][key]["label"], width, style)


def insert_table_cell_image(sheet, row, col, item):
    thumb_path = crawl_handle.get_thubm_path(handle, item["asin"])

    if (thumb_path is None) or (not thumb_path.exists()):
        return

    img = openpyxl.drawing.image.Image(thumb_path)

    # NOTE: マジックナンバー「8」は下記等を参考にして設定．(日本語フォントだと 8 が良さそう)
    # > In all honesty, I cannot tell you how many blogs and stack overflow answers
    # > I read before I stumbled across this magic number: 7.5
    # https://imranhugo.medium.com/how-to-right-align-an-image-in-excel-cell-using-python-and-openpyxl-7ca75a85b13a
    cell_width_pix = TABLE_HEADER["col"]["image"]["width"] * 8
    cell_height_pix = openpyxl.utils.units.points_to_pixels(TABLE_HEADER["row"]["height"])

    cell_width_emu = openpyxl.utils.units.pixels_to_EMU(cell_width_pix)
    cell_height_emu = openpyxl.utils.units.pixels_to_EMU(cell_height_pix)

    margin_pix = 2
    content_width_pix = cell_width_pix - (margin_pix * 2)
    content_height_pix = cell_height_pix - (margin_pix * 2)

    content_ratio = content_width_pix / content_height_pix
    image_ratio = img.width / img.height

    if (img.width > content_width_pix) or (img.height > content_height_pix):
        if image_ratio > content_ratio:
            # NOTE: 画像の横幅をセルの横幅に合わせる
            scale = content_width_pix / img.width
        else:
            # NOTE: 画像の高さをセルの高さに合わせる
            scale = content_height_pix / img.height

        img.width *= scale
        img.height *= scale

    image_width_emu = openpyxl.utils.units.pixels_to_EMU(img.width)
    image_height_emu = openpyxl.utils.units.pixels_to_EMU(img.height)

    col_offset_emu = (cell_width_emu - image_width_emu) / 2
    row_offset_emu = (cell_height_emu - image_height_emu) / 2

    marker = openpyxl.drawing.spreadsheet_drawing.AnchorMarker(
        col=col - 1, row=row - 1, colOff=col_offset_emu, rowOff=row_offset_emu
    )
    marker2 = openpyxl.drawing.spreadsheet_drawing.AnchorMarker(
        col=col, row=row, colOff=-col_offset_emu, rowOff=-row_offset_emu
    )
    size = openpyxl.drawing.xdr.XDRPositiveSize2D(image_width_emu, image_height_emu)

    img.anchor = openpyxl.drawing.spreadsheet_drawing.TwoCellAnchor(_from=marker, to=marker2)

    # img.anchor = openpyxl.drawing.spreadsheet_drawing.OneCellAnchor(_from=marker, ext=size)
    sheet.add_image(img)


def gen_item_cell_style(base_style, key):
    style = base_style.copy()

    if "format" in TABLE_HEADER["col"][key]:
        style["text_format"] = TABLE_HEADER["col"][key]["format"]

    if "wrap" in TABLE_HEADER["col"][key]:
        style["text_wrap"] = TABLE_HEADER["col"][key]["wrap"]
    else:
        style["text_wrap"] = False

    return style


def insert_table_item(sheet, row, item, style):
    for key in TABLE_HEADER["col"].keys():
        col = TABLE_HEADER["col"][key]["pos"]

        cell_style = gen_item_cell_style(style, key)

        if key == "category":
            for i in range(TABLE_HEADER["col"][key]["length"]):
                if i < len(item["category"]):
                    value = item[key][i]
                else:
                    value = ""
                set_item_cell_style(sheet, row, col + i, value, cell_style)
        elif key == "image":
            sheet.cell(row, col).border = cell_style["border"]
            insert_table_cell_image(sheet, row, col, item)
        else:
            set_item_cell_style(sheet, row, col, item[key], cell_style)

        if key == "asin":
            sheet.cell(row, col).hyperlink = item["url"]
        if key == "no":
            sheet.cell(row, col).hyperlink = gen_order_url(item["no"])


def setting_table_view(sheet, row_last):
    crawl_handle.set_status(handle, "テーブルの表示設定しています...")

    sheet.freeze_panes = gen_text_pos(TABLE_HEADER["row"]["pos"] + 1, TABLE_HEADER["col"]["count"]["pos"])

    sheet.auto_filter.ref = "{start}:{end}".format(
        start=gen_text_pos(
            TABLE_HEADER["row"]["pos"], min(map(lambda x: x["pos"], TABLE_HEADER["col"].values()))
        ),
        end=gen_text_pos(row_last, max(map(lambda x: x["pos"], TABLE_HEADER["col"].values()))),
    )
    sheet.sheet_view.showGridLines = False


def insert_sum_row(sheet, row_last, style):
    logging.info("Insert sum row")

    crawl_handle.set_status(handle, "集計行を挿入しています...")

    col = TABLE_HEADER["col"]["price"]["pos"]
    set_item_cell_style(
        sheet,
        row_last + 1,
        col,
        "=sum({cell_first}:{cell_last})".format(
            cell_first=gen_text_pos(TABLE_HEADER["row"]["pos"] + 1, col),
            cell_last=gen_text_pos(row_last, col),
        ),
        gen_item_cell_style(style, "price"),
    )


def generate_list_sheet(book):
    sheet = book.active
    sheet.title = TABLE_SHEET_TITLE

    side = openpyxl.styles.Side(border_style="thin", color="000000")
    border = openpyxl.styles.Border(top=side, left=side, right=side, bottom=side)
    fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="F2F2F2")

    style = {"border": border, "fill": fill}

    row = TABLE_HEADER["row"]["pos"]
    insert_table_header(sheet, row, style)

    item_list = crawl_handle.get_item_list(handle)

    crawl_handle.set_progress_bar(handle, "Insert item", len(item_list))
    crawl_handle.set_status(handle, "購入商品の記載をしています...")

    row += 1
    for item in item_list:
        sheet.row_dimensions[row].height = TABLE_HEADER["row"]["height"]
        insert_table_item(sheet, row, item, style)
        crawl_handle.get_progress_bar(handle, "Insert item").update()
        row += 1

    row_last = row - 1

    crawl_handle.get_progress_bar(handle, "Insert item").update()

    insert_sum_row(sheet, row_last, style)
    setting_table_view(sheet, row_last)


def generate_table_excel(handle, excel_file):
    logging.info("Start to Generate excel file")

    crawl_handle.set_status(handle, "エクセルファイルの作成を開始します...")

    book = openpyxl.Workbook()
    book._named_styles["Normal"].font = openpyxl.styles.Font(name="A-OTF UD新ゴ Pro R", size=12)

    generate_list_sheet(book)

    crawl_handle.set_status(handle, "エクセルファイルを書き出しています...")

    book.save(excel_file)
    book.close()

    logging.info("Complete to Generate excel file")


if __name__ == "__main__":
    import logger
    from config import load_config
    from docopt import docopt

    args = docopt(__doc__)

    logger.init("test", level=logging.INFO)

    config = load_config(args["-c"])
    excel_file = args["-o"]

    handle = crawl_handle.create(config)

    generate_table_excel(handle, excel_file)

    crawl_handle.finish(handle)
